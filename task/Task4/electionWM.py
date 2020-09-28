import openpyxl
wb=openpyxl.load_workbook("Elections.xlsx")
nwb=openpyxl.Workbook()
sheet=nwb["Sheet"]
nwb.remove(sheet)
for sheetname in wb.sheetnames:
    sheet=wb[sheetname]
    rows = sheet.max_row
    new_sheet=nwb.create_sheet(sheetname)
    for data in range(rows):
        w_name=sheet.cell(row=data+1,column=3).value
        r_name=sheet.cell(row=data+1,column=7).value
        w_votes=int(sheet.cell(row=data+1,column=6).value)
        r_votes=int(sheet.cell(row=data+1,column=10).value)
        margin=w_votes - r_votes
        new_sheet.cell(row=data+1,column=1).value=w_name
        new_sheet.cell(row=data+1,column=2).value=r_name
        new_sheet.cell(row=data+1,column=3).value=margin
nwb.save("Electionsmodify.xlsx")
