import openpyxl
import statistics

wb = openpyxl.load_workbook("cricket.xlsx")

new_wb = openpyxl.Workbook()
sheet= new_wb["Sheet"]
new_wb.remove(sheet)

new_wb1= openpyxl.Workbook()
sheet1= new_wb1["Sheet"]
new_wb1.remove(sheet1)

for sheet in  wb.sheetnames:
    name= wb[sheet]
    max_r = name.max_row
    strike_rates = []
   
    for row_line in range(1, max_r):
        runs = name.cell(row = row_line, column = 1).value
        balls = name.cell(row = row_line, column = 2).value
        strike_rate = int(runs/balls*100)
        strike_rates.append(strike_rate)
    sdv = statistics.stdev(strike_rates)
    
    new_sheet1=new_wb.create_sheet(sheet)
    new_sheet2=new_wb1.create_sheet(sheet)
    
    
    for l in strike_rates:
        if l < sdv:
            new_sheet1.cell(row=1, column=1).value=l
            count1+=1
        else:
            new_sheet2.cell(row=1 column=1).value=l
            count2+=1
            
new_wb.save("AboveSTD.xlsx")
new_wb1.save("BelowSTD.xlsx")         
        
                    
