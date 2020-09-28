import openpyxl
wb=openpyxl.load_workbook("Elections.xlsx")
years={}
for sheetname in wb.sheetnames:
    sheet=wb[sheetname]
    rows = sheet.max_row
    margins=[]
    for data in range(rows):
        w_name=sheet.cell(row=data+1,column=3).value
        w_votes=int(sheet.cell(row=data+1,column=6).value)
        r_votes=int(sheet.cell(row=data+1,column=10).value)
        margin=w_votes - r_votes
        margins.append((margin,w_name))
        margins.sort()
    years[sheetname]=margins[-1]
print(years)
        
        
        
