import openpyxl
class Elections:
    def __init__(self,years,sheetname,sheet,rows):
        self.years=years
        self.sheetname=sheetname
        self.
        
    def partiesWon():
            wb=openpyxl.load_workbook("Elections.xlsx")
            years={}
            for sheetname in wb.sheetnames:
                sheet=wb[sheetname]
                rows=sheet.max_row
                parties={}
                for data in range(rows):
                    party=sheet.cell(row=data+1,column=5).value
                    if party in parties:
                        parties[party]+=1
                    else:
                        parties[party]=1
                years[sheetname]=parties
            print(years)
    def margin():
        wb=openpyxl.load_workbook("Elections.xlsx")
        years={}
        for sheetname in wb.sheetnames:
            sheet=wb[sheetname]
            rows = sheet.max_row
            margins=[]
            for data in range(rows):
                w_name=sheet.cell(row=data+1,column=3).value
                w_votes=int(sheet.cell(row=data+1,column=6).value)
                r_votes=int(sheet.cell(row=data+1,column=10).value)
                margin=w_votes - r_votes
                margins.append((margin,w_name))
                margins.sort()
            years[sheetname]=margins[-1]
        print(years)

    def femaleCand():
        wb=openpyxl.load_workbook("Elections.xlsx")
        years={}
        for sheetname in wb.sheetnames:
            sheet=wb[sheetname]
            rows = sheet.max_row
            f_cand=0
            f_win=0
                
            for data in range(rows):
                w_gen=sheet.cell(row=data+1,column=4).value
                r_gen=sheet.cell(row=data+1,column=8).value

                if w_gen == 'F':
                     f_cand+=1
                     f_win+=1
                if r_gen == 'F':
                    f_cand+=1
                years[sheetname]={"Female candidates":f_cand,"Winners":f_win}
        print(years)
