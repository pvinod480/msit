import openpyxl
wb=openpyxl.load_workbook("Elections.xlsx")
years={}
for sheetname in wb.sheetnames:
    sheet=wb[sheetname]
    rows = sheet.max_row
    f_cand=0
    f_win=0
        
    for data in range(rows):
        w_gen=sheet.cell(row=data+1,column=4).value
        r_gen=sheet.cell(row=data+1,column=8).value

        if w_gen == 'F':
             f_cand+=1
             f_win+=1
        if r_gen == 'F':
            f_cand+=1
        years[sheetname]={"Female candidates":f_cand,"Winners":f_win}
print(years)

             
        
    
