import openpyxl
wb=openpyxl.load_workbook("Elections.xlsx")
years={}
for sheetname in wb.sheetnames:
    sheet=wb[sheetname]
    rows=sheet.max_row
    parties={}
    for data in range(rows):
        party=sheet.cell(row=data+1,column=5).value
        if party in parties:
            parties[party]+=1
        else:
            parties[party]=1
    years[sheetname]=parties
print(years)

